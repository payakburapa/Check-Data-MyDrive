import datetime
from openpyxl.styles import Font
from openpyxl.styles.colors import Color
from datetime import datetime
from datetime import date
import os
import openpyxl
import array as arr
from openpyxl import Workbook


def headder(sheet):
    dc = openpyxl.load_workbook("MyDrive.xlsx")
    sh = dc[sheet]
    sh.row_dimensions[1].height = 30
    sh.column_dimensions['A'].width = 19
    sh.column_dimensions['B'].width = 17.43
    sh.column_dimensions['C'].width = 11.29
    sh.column_dimensions['D'].width = 8.43
    sh.column_dimensions['E'].width = 12
    sh.column_dimensions['F'].width = 26.29
    sh.column_dimensions['G'].width = 8.43
    sh.cell(row=1, column=1).value = "Date\n(YYYY-MM-DD)"
    sh.cell(row=1, column=2).value = "Duration(Day)"
    sh.cell(row=1, column=3).value = "Time"
    sh.cell(row=1, column=4).value = "Type"
    sh.cell(row=1, column=5).value = "Size(bytes)"
    sh.cell(row=1, column=6).value = "Name"
    sh.cell(row=1, column=7).value = "Path"
    dc.save(filename = 'MyDrive.xlsx')


def writedata(sht,data,linecount,path):    
    CELL_1 = int(linecount)
    excel = openpyxl.load_workbook('MyDrive.xlsx')
    excel.active
    CELL_1 += 1
    CA = 'A' + str(CELL_1)
    CB = 'B' + str(CELL_1)
    CC = 'C' + str(CELL_1)
    CD = 'D' + str(CELL_1)
    CE = 'E' + str(CELL_1)
    CF = 'F' + str(CELL_1)
    CG = 'G' + str(CELL_1)

    if data.split(" ")[0:][0] != "":
        
        Date = datetime.strptime(data.split(" ")[0:][0],'%m/%d/%Y')
        excel[sht][CA] = Date.date() ## Date file
        daynow = date.today()
        datefile = Date.date()
        count = daynow - datefile
        countdate = str(count).split(" ")[0:][0]
        excel[sht][CB] = countdate ## Count Duration(Day)
        excel[sht][CC] = data.split(" ")[0:][2] + " " + data.split(" ")[0:][3] ## Time
        
        if data.split(" ")[0:][7] != "":
            excel[sht][CD] = data.split(" ")[0:][7]
        else:
            excel[sht][CD] = "File"

        if data.split(" ")[0:][3] != "" and data[20:].strip().split(" ")[0:][0] != "<DIR>":
            excel[sht][CE] = data[20:].strip().split(" ")[0:][0]
        else:
            excel[sht][CE] = ""

        if data.split(" ")[0:][3] != "":
            excel[sht][CF] = data[20:].strip().split(" ")[-1:][0]
        else:
            excel[sht][CF] = ""

        excel[sht][CG] = path
        print(path)
    else:
        excel[sht][CA] = ""
        excel[sht][CB] = ""
        excel[sht][CC] = data.strip().split(" ")[0:][0] + " " + data.strip().split(" ")[1:][0]
        excel[sht][CE] = data.strip().split(" ")[-2:][0]
        excel[sht][CF] = "bytes"
        excel[sht][CG] = path

    excel.save(filename = 'MyDrive.xlsx')
    
    
f = open("MyDrive.txt","r")
Lines = f.readlines()
count = 0
c = 0
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("inetpub")
ws2 = wb.create_sheet("Intel")
ws3 = wb.create_sheet("KBSERVICE.SHUTDOWN")
ws4 = wb.create_sheet("PerfLogs")
ws5 = wb.create_sheet("Program Files")
ws6 = wb.create_sheet("Program (x86)")
ws7 = wb.create_sheet("StartUSB")
ws8 = wb.create_sheet("Users")
ws9 = wb.create_sheet("Windows")
wb.save(filename = 'MyDrive.xlsx')

sh = ''
shc = ''
path = ''
countin = 0

os.system('dir /s C:\ > D:\\Users\\XXXXXX\\Desktop\\Python-Test\\MyDrive.txt')

for line in Lines:
    count += 1
    if count >= 6 and count <= 14:
        if line.lstrip():
            wb = Workbook()
            ws = wb.active
            sht = line.lstrip()
            if sht.split()[4:][0] != sht.split()[4:][-1]:
                result=sht.split()[4:][0] + " " + sht.split()[4:][-1]
            else:
                result=sht.split()[4:][0]
                
            headder(result)

    linestr = line.replace('\n','')
    if linestr!="" and count!=4 and count >= 16:
        if linestr.split(" ")[1:][0]=="Directory":
            sh=linestr[1:500].split("\\")[1:][0]## get sheet
            path=linestr[0:500]## get path

        if linestr.split(" ")[1:][0]!="Directory":
            if sh == shc:
                countin += 1
            else:
                countin = 1
                shc = sh
                
            writedata(sh,linestr,str(countin),path)

